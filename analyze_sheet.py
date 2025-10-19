#!/usr/bin/env python3
"""
Analyze Excel sheet structure for Long Khánh Confessions
"""
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('Long Khánh Confessions (Câu trả lời).xlsx')
ws = wb.active

# Get headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)

# Print all column headers with their indices
print("=" * 80)
print("COLUMN STRUCTURE OF YOUR CONFESSIONS SHEET")
print("=" * 80)
for idx, header in enumerate(headers, start=1):
    col_letter = chr(64 + idx) if idx <= 26 else f"{chr(64 + (idx-1)//26)}{chr(65 + (idx-1)%26)}"
    print(f"Column {col_letter} (#{idx}): {header}")

# Show sample data from first 3 rows for key columns
print("\n" + "=" * 80)
print("SAMPLE DATA FROM FIRST 3 CONFESSIONS")
print("=" * 80)

for row_num in range(2, min(5, ws.max_row + 1)):
    print(f"\n--- Row {row_num - 1} ---")
    for idx, header in enumerate(headers, start=1):
        cell_value = ws.cell(row=row_num, column=idx).value
        if cell_value and str(cell_value).strip():  # Only show non-empty cells
            col_letter = chr(64 + idx) if idx <= 26 else f"{chr(64 + (idx-1)//26)}{chr(65 + (idx-1)%26)}"
            # Truncate long text
            display_value = str(cell_value)[:100] + "..." if len(str(cell_value)) > 100 else str(cell_value)
            print(f"  {col_letter}. {header}: {display_value}")

print("\n" + "=" * 80)
print(f"Total rows with data: {ws.max_row - 1}")
print("=" * 80)
