#!/usr/bin/env python3
"""
Analyze which columns have actual confession content
"""
from openpyxl import load_workbook

wb = load_workbook('Long Khánh Confessions (Câu trả lời).xlsx')
ws = wb.active

# Columns of interest for text content
content_columns = {
    'E': 5, 'F': 6, 'G': 7, 'H': 8, 'L': 12, 'M': 13
}

# Count non-empty cells in each column
print("=" * 80)
print("CONTENT ANALYSIS - Which columns have confession data?")
print("=" * 80)

for col_letter, col_idx in content_columns.items():
    header = ws.cell(row=1, column=col_idx).value
    non_empty = 0
    sample_texts = []

    for row_num in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=col_idx).value
        if cell_value and str(cell_value).strip():
            non_empty += 1
            if len(sample_texts) < 2:  # Get 2 samples
                sample_texts.append(str(cell_value)[:80] + "...")

    print(f"\nColumn {col_letter}: {header}")
    print(f"  - Non-empty entries: {non_empty} / {ws.max_row - 1}")
    print(f"  - Percentage: {(non_empty / (ws.max_row - 1) * 100):.1f}%")
    if sample_texts:
        print(f"  - Sample 1: {sample_texts[0]}")
        if len(sample_texts) > 1:
            print(f"  - Sample 2: {sample_texts[1]}")

print("\n" + "=" * 80)
print("RECOMMENDATION FOR DATA EXTRACTION")
print("=" * 80)
print("Based on the analysis above, we should combine text from:")
print("- Columns with >50% data for comprehensive analysis")
print("- This gives us the most complete picture of student confessions")
